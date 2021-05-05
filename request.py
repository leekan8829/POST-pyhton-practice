import json
from openpyxl import Workbook
# 預設可讀寫，若有需要可以指定write_only和read_only為True
import requests
import time
from mail_auto import sendmail

url = "Your api url"
'''-------------------------------------------------------------------------------------------------'''

#def move_fan(old_meter_id,new_meter_id,old_meter_new_fan_id):
def move_fan(*info):
    url = "Your api url"
    count=0
    count = len(info)
    if(count == 2):
        body = {
            "old_meter_id": info[0] ,
            "new_meter_id": info[1]
        }
    else:
        body = {
            "old_meter_id": info[0],  
            "new_meter_id": info[1],
            "old_meter_new_fan": info[2]
        }
    headers = {
        "Authorization":"Your Access Token"
    }
    resp = requests.post(url,headers=headers, json=body)
    resp_list = resp.json()
    #type is dict
    return resp_list

def multi_move_fans2(old_meter_list,new_meter_list,old_meter_new_fan_list):
    list_to_deal = []
    list_all_response = []
    for x,y,z in zip(old_meter_list,new_meter_list,old_meter_new_fan_list):
    #將要處理的old_meter_list,new_meter_list,old_meter_new_fan_list 為一個set放進 list_to_deal
        temp = [x,y,z]
        list_to_deal.append(temp)
    print('list_to_deal:')
    print(list_to_deal)
    print('\n\n')
    for x in list_to_deal:
    #開始處理要打endpoint的資料
        list_parm = []
        list_response = []
        #list_parm放要打endpoint的參數
        for y in x:
            list_parm.append(y)
        list_response = move_fan(list_parm[0],list_parm[1],list_parm[2])
        list1 = list_response['old_meter']
        list2 = list_response['new_meter']
        #將每一個打完endpoint的response放進list_all_response
        rows = [
            ['old_meter_id','old_fan_id','old_comment','new_meter_id','new_fan_id','new_comment'],
            [list1['meter_id'],list1['fan_id'],list1['comment'],list2['meter_id'],list2['fan_id'],list2['comment']],
            [' ',' ',' ']
        ]
        list_all_response.append(rows)
        list_parm.clear()
        #打完一次endpoint之後清空list_parm
    return list_all_response

def send_movefan_report(list_all_response):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'report'
    for rows in list_all_response:
        for row in rows:
            sheet1.append(row)
            print(row)
    timestr=time.strftime("%m%d%H%M")
    newfilename = ('%sMove_fan_report.xlsx'%timestr)
    wb.save(newfilename)
    sendmail('move fan report',newfilename)




def multi_move_fans(old_meter_list,new_meter_list,old_meter_new_fan_list):
    list_to_deal = []
    list_all_response = []
    for x,y,z in zip(old_meter_list,new_meter_list,old_meter_new_fan_list):
    #將要處理的old_meter_list,new_meter_list,old_meter_new_fan_list 為一個set放進 list_to_deal
        temp = [x,y,z]
        list_to_deal.append(temp)
    print('list_to_deal:')
    print(list_to_deal)
    print('\n\n')
    for x in list_to_deal:
    #開始處理要打endpoint的資料
        list_parm = []
        #list_parm放要打endpoint的參數
        for y in x:
            list_parm.append(y)
        list_all_response.append(move_fan(list_parm[0],list_parm[1],list_parm[2]) )
        #將每一個打完endpoint的response放進list_all_response
        print('complete')
        print(list_parm)
        print('\n\n')
        list_parm.clear()
        #打完一次endpoint之後清空list_parm
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'report'
    for rows in list_all_response:
        for row in rows:
            sheet1.append(row)
            print(row)
    timestr=time.strftime("%m%d%H%M")
    #wb.save(r'%sexample.xlsx'%timestr)
    newfilename = ('%sexample.xlsx'%timestr)
    wb.save(newfilename)
    sendmail('move fan report',newfilename)

'''-------------------------------------------------------------------------------------------------'''
#換表不換fan
def replace_meter(old_meter_id,new_meter_id):
    url = "Your api url"
    body = {
        "old_meter_id": old_meter_id,  
        "new_meter_id": new_meter_id
    }
    headers = {
        "Authorization":"Acess Token"
    }
    resp = requests.post(url,headers=headers, json=body)
    resp_list = resp.json()
    # print(type(resp_list))
    # print(resp.json())
    list1 = resp_list['old_meter']
    list2 = resp_list['new_meter']
    #去拿要寫在excel的元素
    print('old_meter: '+list1['meter_id']+'  fan_id: '+list1['fan_id']+'  old_meter_comment: '+list1['comment'])
    print('new_meter: '+list2['meter_id']+'  fan_id: '+list2['fan_id']+'  new_meter_comment: '+list1['comment'])
    print('\n')
    rows = [
        ['meter_id','fan_id','comment'],
        [list1['meter_id'],list1['fan_id'],list1['comment']],
        [list2['meter_id'],list2['fan_id'],list2['comment']],
        [' ',' ',' ']
    ]
    print(rows)
    return rows

def multi_replace_meter(old_meter_list,new_meter_list):
    list_to_deal = []
    list_all_response = []
    for x,y in zip(old_meter_list,new_meter_list):
        temp = [x,y]
        list_to_deal.append(temp)
    print('list_to_deal:')
    print(list_to_deal)
    print('\n\n')
    for x in list_to_deal:
    #開始處理要打endpoint的資料
        list_parm = []
        #list_parm放要打endpoint的參數
        for y in x:
            list_parm.append(y)
        list_all_response.append(replace_meter(list_parm[0],list_parm[1]))
        #將每一個打完endpoint的response放進list_all_response
        print('complete')
        print(list_parm)
        print('\n\n')
        list_parm.clear()
        #打完一次endpoint之後清空list_parm
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'report'
    for rows in list_all_response:
        for row in rows:
            sheet1.append(row)
            print(row)
    timestr=time.strftime("%m%d%H%M")
    #wb.save(r'%sexample.xlsx'%timestr)
    newfilename = ('%sexample.xlsx'%timestr)
    wb.save(newfilename)
    sendmail('replace meter report',newfilename)

'''-------------------------------------------------------------------------------------------------'''


if __name__ == "__main__":
    multi_move_fans(old_meter_list1,new_meter_list1,old_meter_new_fan_list1)


